import io

import pytest
from openpyxl import Workbook, load_workbook

from agente_oracle.tools.ferramentas.juntar_excel import ArquivoExcelInvalido, analisar_colunas, juntar_planilhas


def _bytes_planilha(cabecalho: list[str], linhas: list[list]) -> bytes:
    workbook = Workbook()
    planilha = workbook.active
    planilha.append(cabecalho)
    for linha in linhas:
        planilha.append(linha)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


class TestJuntarPlanilhas:
    def test_colunas_iguais_empilha_num_bloco_so(self) -> None:
        arquivo1 = _bytes_planilha(["nome", "idade"], [["Ana", 30]])
        arquivo2 = _bytes_planilha(["nome", "idade"], [["Bruno", 25]])

        resultado = load_workbook(io.BytesIO(juntar_planilhas(arquivo1, arquivo2)))
        planilha = resultado.active
        linhas = list(planilha.iter_rows(values_only=True))

        assert linhas == [("nome", "idade"), ("Ana", 30), ("Bruno", 25)]

    def test_colunas_iguais_em_ordem_diferente_reordena_segunda_planilha(self) -> None:
        arquivo1 = _bytes_planilha(["nome", "idade"], [["Ana", 30]])
        arquivo2 = _bytes_planilha(["idade", "nome"], [[25, "Bruno"]])

        resultado = load_workbook(io.BytesIO(juntar_planilhas(arquivo1, arquivo2)))
        linhas = list(resultado.active.iter_rows(values_only=True))

        assert linhas == [("nome", "idade"), ("Ana", 30), ("Bruno", 25)]

    def test_colunas_iguais_nao_aplica_cor_de_fundo(self) -> None:
        arquivo1 = _bytes_planilha(["nome"], [["Ana"]])
        arquivo2 = _bytes_planilha(["nome"], [["Bruno"]])

        resultado = load_workbook(io.BytesIO(juntar_planilhas(arquivo1, arquivo2)))
        planilha = resultado.active

        for linha in planilha.iter_rows():
            for celula in linha:
                assert celula.fill.fgColor.rgb in (None, "00000000")

    def test_colunas_diferentes_mantem_dois_blocos_lado_a_lado_com_cores(self) -> None:
        arquivo1 = _bytes_planilha(["nome"], [["Ana"]])
        arquivo2 = _bytes_planilha(["produto"], [["Caneta"]])

        resultado = load_workbook(io.BytesIO(juntar_planilhas(arquivo1, arquivo2)))
        planilha = resultado.active
        linhas = list(planilha.iter_rows(values_only=True))

        # bloco 1 na coluna A, coluna B em branco (separação), bloco 2 na coluna C
        assert linhas == [("nome", None, "produto"), ("Ana", None, "Caneta")]

        cabecalho1 = planilha["A1"]
        dado1 = planilha["A2"]
        cabecalho2 = planilha["C1"]
        dado2 = planilha["C2"]

        assert cabecalho1.fill.fgColor.rgb == "00D9F2D9"
        assert dado1.fill.fgColor.rgb == "00D9F2D9"
        assert cabecalho2.fill.fgColor.rgb == "00FCE4D6"
        assert dado2.fill.fgColor.rgb == "00FCE4D6"

    def test_arquivo_invalido_levanta_erro(self) -> None:
        arquivo_valido = _bytes_planilha(["nome"], [["Ana"]])

        with pytest.raises(ArquivoExcelInvalido):
            juntar_planilhas(b"isso nao e um xlsx", arquivo_valido)

    def test_linha_totalmente_em_branco_e_descartada(self) -> None:
        arquivo1 = _bytes_planilha(["nome", "idade"], [["Ana", 30], [None, None]])
        arquivo2 = _bytes_planilha(["nome", "idade"], [["Bruno", 25]])

        resultado = load_workbook(io.BytesIO(juntar_planilhas(arquivo1, arquivo2)))
        linhas = list(resultado.active.iter_rows(values_only=True))

        assert linhas == [("nome", "idade"), ("Ana", 30), ("Bruno", 25)]


class TestJuntarPlanilhasPorChaveComum:
    def test_coluna_comum_com_correspondencia_simples_junta_numa_linha_so(self) -> None:
        arquivo1 = _bytes_planilha(["filial", "nome"], [["01", "Fazenda A"]])
        arquivo2 = _bytes_planilha(["filial", "valor"], [["01", 100]])

        resultado = load_workbook(io.BytesIO(juntar_planilhas(arquivo1, arquivo2)))
        linhas = list(resultado.active.iter_rows(values_only=True))

        assert linhas == [("filial", "nome", "valor"), ("01", "Fazenda A", 100)]

    def test_valor_repetido_de_um_lado_multiplica_as_linhas(self) -> None:
        arquivo1 = _bytes_planilha(["filial", "nome"], [["01", "Fazenda A"]])
        arquivo2 = _bytes_planilha(["filial", "valor"], [["01", 100], ["01", 200]])

        resultado = load_workbook(io.BytesIO(juntar_planilhas(arquivo1, arquivo2)))
        linhas = list(resultado.active.iter_rows(values_only=True))

        assert linhas == [
            ("filial", "nome", "valor"),
            ("01", "Fazenda A", 100),
            ("01", "Fazenda A", 200),
        ]

    def test_sem_correspondencia_mantem_linha_com_colunas_em_branco(self) -> None:
        arquivo1 = _bytes_planilha(["filial", "nome"], [["01", "Fazenda A"], ["02", "Fazenda B"]])
        arquivo2 = _bytes_planilha(["filial", "valor"], [["01", 100], ["03", 300]])

        resultado = load_workbook(io.BytesIO(juntar_planilhas(arquivo1, arquivo2)))
        linhas = list(resultado.active.iter_rows(values_only=True))

        assert linhas == [
            ("filial", "nome", "valor"),
            ("01", "Fazenda A", 100),
            ("02", "Fazenda B", None),
            ("03", None, 300),
        ]

    def test_resultado_da_juncao_sai_inteiro_verde(self) -> None:
        arquivo1 = _bytes_planilha(["filial", "nome"], [["01", "Fazenda A"]])
        arquivo2 = _bytes_planilha(["filial", "valor"], [["01", 100]])

        resultado = load_workbook(io.BytesIO(juntar_planilhas(arquivo1, arquivo2)))
        planilha = resultado.active

        for linha in planilha.iter_rows():
            for celula in linha:
                assert celula.fill.fgColor.rgb == "00D9F2D9"

    def test_linha_fantasma_so_com_a_chave_e_descartada_antes_da_juncao(self) -> None:
        arquivo1 = _bytes_planilha(
            ["filial", "nome"],
            [["01", "Fazenda A"], ["01", None]],  # 2ª linha: só a chave, sem dado próprio
        )
        arquivo2 = _bytes_planilha(["filial", "valor"], [["01", 100]])

        resultado = load_workbook(io.BytesIO(juntar_planilhas(arquivo1, arquivo2)))
        linhas = list(resultado.active.iter_rows(values_only=True))

        assert linhas == [("filial", "nome", "valor"), ("01", "Fazenda A", 100)]


class TestAnalisarColunas:
    def test_colunas_identicas(self) -> None:
        arquivo1 = _bytes_planilha(["nome", "idade"], [["Ana", 30]])
        arquivo2 = _bytes_planilha(["idade", "nome"], [[25, "Bruno"]])

        analise = analisar_colunas(arquivo1, arquivo2)

        assert analise["tipo"] == "identicas"
        assert analise["colunas1"] == ["nome", "idade"]
        assert analise["colunas2"] == ["idade", "nome"]
        assert set(analise["colunas_comuns"]) == {"nome", "idade"}

    def test_colunas_parcialmente_em_comum(self) -> None:
        arquivo1 = _bytes_planilha(["filial", "nome"], [["01", "Fazenda A"]])
        arquivo2 = _bytes_planilha(["filial", "valor"], [["01", 100]])

        analise = analisar_colunas(arquivo1, arquivo2)

        assert analise["tipo"] == "parcial"
        assert analise["colunas_comuns"] == ["filial"]

    def test_nenhuma_coluna_em_comum(self) -> None:
        arquivo1 = _bytes_planilha(["nome"], [["Ana"]])
        arquivo2 = _bytes_planilha(["produto"], [["Caneta"]])

        analise = analisar_colunas(arquivo1, arquivo2)

        assert analise["tipo"] == "nenhuma"
        assert analise["colunas_comuns"] == []

    def test_arquivo_invalido_levanta_erro(self) -> None:
        arquivo_valido = _bytes_planilha(["nome"], [["Ana"]])

        with pytest.raises(ArquivoExcelInvalido):
            analisar_colunas(b"isso nao e um xlsx", arquivo_valido)
