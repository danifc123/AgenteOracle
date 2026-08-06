"""Lógica pura (sem HTTP) de junção de duas planilhas `.xlsx` — ver
`server/ferramentas/juntar_excel.py` para a rota que expõe isso via upload.
Regra de negócio, em ordem de prioridade:
1. Colunas idênticas (mesmo conjunto de nomes, ordem não importa) -> empilha
   tudo num bloco só, cabeçalho único.
2. Pelo menos uma coluna em comum, mas não todas -> JOIN completo (outer)
   pelas colunas em comum (`_juntar_por_chave_comum`): a coluna comum aparece
   uma vez só, e cada linha de um lado combina com TODA linha do outro lado
   que tenha o mesmo valor nessa coluna (se um dos lados tiver mais de uma
   linha com o mesmo valor, cada combinação vira uma linha — mesmo
   comportamento de JOIN de banco de dados). Linha sem correspondência do
   outro lado entra do mesmo jeito, com as colunas que faltam em branco. O
   resultado inteiro sai verde clarinho (mesmo tom do bloco único do caso 1)
   — é uma junção bem-sucedida, não faz sentido diferenciar origem por cor.
3. Nenhuma coluna em comum -> não dá pra combinar de verdade (nenhuma
   junção encontrada), os dois blocos ficam lado a lado (uma coluna em
   branco entre eles), verde clarinho pro 1º arquivo e laranja clarinho pro
   2º — únicas cores usadas no preview do frontend (`D9F2D9`/`FCE4D6`).

Linhas "fantasma" (só com valor na coluna-chave, todo o resto em branco —
sobra comum de planilha exportada com range usado maior que os dados reais)
são descartadas antes do JOIN: não têm dado nenhum pra contribuir e só
inflam o resultado com combinações sem sentido."""

import io

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

_VERDE_CLARO = PatternFill(start_color="D9F2D9", end_color="D9F2D9", fill_type="solid")
_LARANJA_CLARO = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
_LARGURA_MAXIMA_COLUNA = 50


class ArquivoExcelInvalido(Exception):
    """Levantada quando um dos arquivos enviados não é um .xlsx válido."""


def _ler_planilha(conteudo: bytes) -> tuple[list[str], list[list]]:
    """Lê a planilha ATIVA de um .xlsx: 1ª linha vira cabeçalho, o resto vira
    linhas de dado. `data_only=True` lê o valor calculado de fórmula (não a
    fórmula em si) — mais previsível pra planilha de origem desconhecida."""
    try:
        workbook = load_workbook(io.BytesIO(conteudo), data_only=True)
    except Exception as erro:
        raise ArquivoExcelInvalido(f"Não foi possível ler o arquivo: {erro}") from erro

    linhas = list(workbook.active.iter_rows(values_only=True))
    if not linhas:
        return [], []

    cabecalho = [str(valor) if valor is not None else "" for valor in linhas[0]]
    # `iter_rows` segue o "range usado" da planilha, que costuma ir além dos
    # dados reais (formatação aplicada a mais) — descarta linha totalmente
    # em branco pra não virar dado fantasma no resultado.
    linhas_de_dados = [list(linha) for linha in linhas[1:] if any(valor is not None for valor in linha)]
    return cabecalho, linhas_de_dados


def _escrever_bloco(
    planilha,
    cabecalho: list[str],
    linhas: list[list],
    preenchimento: PatternFill | list[PatternFill | None] | None,
    coluna_inicial: int = 1,
) -> None:
    """Escreve um bloco (cabeçalho em negrito + linhas) a partir da linha 1,
    começando em `coluna_inicial` — permite colocar dois blocos lado a lado
    na mesma planilha em vez de um embaixo do outro. `preenchimento` pode ser
    uma cor só (aplicada em todas as colunas do bloco) ou uma lista de cores,
    uma por coluna (usado no resultado do JOIN, onde cada coluna pode ter uma
    origem diferente)."""
    cores = preenchimento if isinstance(preenchimento, list) else [preenchimento] * len(cabecalho)

    for indice_coluna, (valor, cor) in enumerate(zip(cabecalho, cores)):
        celula = planilha.cell(row=1, column=coluna_inicial + indice_coluna, value=valor)
        celula.font = Font(bold=True)
        if cor is not None:
            celula.fill = cor

    for indice_linha, linha in enumerate(linhas, start=2):
        for indice_coluna, (valor, cor) in enumerate(zip(linha, cores)):
            celula = planilha.cell(row=indice_linha, column=coluna_inicial + indice_coluna, value=valor)
            if cor is not None:
                celula.fill = cor


def _juntar_por_chave_comum(
    cabecalho1: list[str],
    linhas1: list[list],
    cabecalho2: list[str],
    linhas2: list[list],
    colunas_comuns: list[str],
) -> tuple[list[str], list[list], list[PatternFill | None]]:
    """JOIN completo (outer) pelas colunas em comum: linhas com o mesmo valor
    nelas viram uma linha só (toda combinação, se um dos lados tiver mais de
    uma linha com o mesmo valor); linhas sem correspondência do outro lado
    entram do mesmo jeito, com as colunas que faltam em branco. Linha
    "fantasma" (só a chave preenchida, resto em branco) é descartada antes —
    não tem dado nenhum pra contribuir e só infla o resultado."""
    indices_comuns_1 = [cabecalho1.index(coluna) for coluna in colunas_comuns]
    indices_comuns_2 = [cabecalho2.index(coluna) for coluna in colunas_comuns]
    indices_unicos_1 = [i for i in range(len(cabecalho1)) if cabecalho1[i] not in colunas_comuns]
    indices_unicos_2 = [i for i in range(len(cabecalho2)) if cabecalho2[i] not in colunas_comuns]

    cabecalho_final = (
        list(colunas_comuns) + [cabecalho1[i] for i in indices_unicos_1] + [cabecalho2[i] for i in indices_unicos_2]
    )
    # Junção bem-sucedida sai inteira verde clarinho — mesmo tom do bloco
    # único do caso de colunas idênticas — sem diferenciar origem por coluna.
    cores_final: list[PatternFill | None] = [_VERDE_CLARO] * len(cabecalho_final)

    def _chave(linha: list, indices: list[int]) -> tuple:
        return tuple(linha[i] for i in indices)

    def _tem_dado_proprio(linha: list, indices_unicos: list[int]) -> bool:
        return any(linha[i] is not None for i in indices_unicos)

    linhas1 = [linha for linha in linhas1 if _tem_dado_proprio(linha, indices_unicos_1)]
    linhas2 = [linha for linha in linhas2 if _tem_dado_proprio(linha, indices_unicos_2)]

    grupos2: dict[tuple, list[list]] = {}
    for linha in linhas2:
        grupos2.setdefault(_chave(linha, indices_comuns_2), []).append(linha)

    chaves_planilha1 = {_chave(linha, indices_comuns_1) for linha in linhas1}

    linhas_final: list[list] = []
    for linha1 in linhas1:
        chave1 = _chave(linha1, indices_comuns_1)
        valores_comuns = list(chave1)
        valores_unicos_1 = [linha1[i] for i in indices_unicos_1]
        correspondentes = grupos2.get(chave1)

        if correspondentes:
            for linha2 in correspondentes:
                valores_unicos_2 = [linha2[i] for i in indices_unicos_2]
                linhas_final.append(valores_comuns + valores_unicos_1 + valores_unicos_2)
        else:
            linhas_final.append(valores_comuns + valores_unicos_1 + [None] * len(indices_unicos_2))

    for chave2, linhas_grupo in grupos2.items():
        if chave2 not in chaves_planilha1:
            for linha2 in linhas_grupo:
                valores_unicos_2 = [linha2[i] for i in indices_unicos_2]
                linhas_final.append(list(chave2) + [None] * len(indices_unicos_1) + valores_unicos_2)

    return cabecalho_final, linhas_final, cores_final


def _aplicar_largura_automatica(planilha) -> None:
    for coluna in planilha.columns:
        maior_valor = max((len(str(celula.value)) for celula in coluna if celula.value is not None), default=0)
        planilha.column_dimensions[coluna[0].column_letter].width = min(maior_valor + 2, _LARGURA_MAXIMA_COLUNA)


def juntar_planilhas(conteudo1: bytes, conteudo2: bytes) -> bytes:
    cabecalho1, linhas1 = _ler_planilha(conteudo1)
    cabecalho2, linhas2 = _ler_planilha(conteudo2)

    workbook = Workbook()
    planilha = workbook.active
    planilha.title = "Planilhas combinadas"

    colunas_comuns = [coluna for coluna in cabecalho1 if coluna in cabecalho2]

    if cabecalho1 and set(cabecalho1) == set(cabecalho2):
        indices = [cabecalho2.index(coluna) for coluna in cabecalho1]
        linhas2_reordenadas = [[linha[indice] for indice in indices] for linha in linhas2]
        _escrever_bloco(planilha, cabecalho1, linhas1 + linhas2_reordenadas, preenchimento=None)
    elif colunas_comuns:
        cabecalho_final, linhas_final, cores_final = _juntar_por_chave_comum(
            cabecalho1, linhas1, cabecalho2, linhas2, colunas_comuns
        )
        _escrever_bloco(planilha, cabecalho_final, linhas_final, cores_final)
    else:
        _escrever_bloco(planilha, cabecalho1, linhas1, _VERDE_CLARO)
        coluna_bloco2 = len(cabecalho1) + 2  # 1 coluna em branco separando os dois blocos
        _escrever_bloco(planilha, cabecalho2, linhas2, _LARANJA_CLARO, coluna_inicial=coluna_bloco2)

    _aplicar_largura_automatica(planilha)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
